from openpyxl import load_workbook

wb = load_workbook(filename = 'spreadsheets/test_datetime.xlsx', read_only=True)
ws = wb["Sheet1"]



for i, row in enumerate(ws.iter_rows()):
	for cell in row:
		print(i, cell.value)